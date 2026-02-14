# Finance Shark Game for MSRUAS - FIXED Streamlit Version
# Run: pip install streamlit pandas numpy openpyxl
# Then: streamlit run this_file.py

import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import io

# Page config
st.set_page_config(
    page_title="Finance Shark Game - MSRUAS",
    page_icon="🦈",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {font-size: 3rem; color: #ff6b6b; text-align: center; margin-bottom: 1rem;}
    .shark-win {background-color: #d4edda; padding: 2rem; border-radius: 1rem; border: 3px solid #28a745;}
    .shark-lose {background-color: #f8d7da; padding: 2rem; border-radius: 1rem; border: 3px solid #dc3545;}
    .metric-card {background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 1.5rem; border-radius: 1rem; text-align: center;}
</style>
""", unsafe_allow_html=True)

# Title
st.markdown('<h1 class="main-header">🦈 Finance Shark Game</h1>', unsafe_allow_html=True)
st.markdown("**MSRUAS Management Fest | Shark Tank India Business Analysis Game**")

# Game Data
@st.cache_data
def load_scenarios():
    return {
        "Tech Startup (Meesho style)": {
            "cashflows": [-5000000, 1000000, 2000000, 3500000, 5000000, 7000000],
            "assets": {"cash": 1500000, "inventory": 2000000, "total": 8500000},
            "liabilities": {"debt": 3000000, "equity": 5500000},
            "net_profit": 1200000
        },
        "Food Delivery (Swiggy style)": {
            "cashflows": [-10000000, 2000000, 3000000, 4500000, 6000000, 8000000],
            "assets": {"cash": 2500000, "inventory": 3500000, "total": 12000000},
            "liabilities": {"debt": 5000000, "equity": 7000000},
            "net_profit": 1800000
        }
    }

scenarios = load_scenarios()

# Sidebar
st.sidebar.title("🎮 Game Controls")
game_mode = st.sidebar.selectbox("Select Mode:", ["Play Game", "Excel Template"], index=0)

if game_mode == "Play Game":
    # Main Game Area
    st.header("📈 Analyze Startup Financials")
    
    col1, col2 = st.columns(2)
    
    with col1:
        scenario_name = st.selectbox("Choose Startup:", list(scenarios.keys()))
        data = scenarios[scenario_name]
        
        # Discount Rate
        discount_rate = st.slider("Discount Rate (%)", 5.0, 25.0, 10.0) / 100
        
        # Cash Flows Table
        years = ["Y0 (Invest)", "Y1", "Y2", "Y3", "Y4", "Y5"]
        cf_df = pd.DataFrame({
            "Year": years,
            "Cash Flow (₹)": [f"₹{x:,}" for x in data["cashflows"]],
            "Discounted CF": [f"₹{x/(1+discount_rate)**i:,.0f}" for i, x in enumerate(data["cashflows"])]
        })
        st.subheader("💰 Cash Flow Analysis")
        st.dataframe(cf_df, use_container_width=True)
        
        # NPV Calculation
        npv_value = np.npv(discount_rate, data["cashflows"])
        st.metric("NPV", f"₹{npv_value:,.0f}", delta=None)
    
    with col2:
        # Balance Sheet
        st.subheader("📊 Balance Sheet (Year 3)")
        bs_data = {
            "Items": ["Cash", "Inventory", "Total Assets", "Debt", "Equity"],
            "Amount (₹)": [
                f"₹{data['assets']['cash']:,}",
                f"₹{data['assets']['inventory']:,}", 
                f"₹{data['assets']['total']:,}",
                f"₹{data['liabilities']['debt']:,}",
                f"₹{data['liabilities']['equity']:,}"
            ]
        }
        st.dataframe(pd.DataFrame(bs_data), use_container_width=True)
    
    # Key Ratios
    st.subheader("🔢 Financial Ratios")
    col1, col2, col3 = st.columns(3)
    
    current_ratio = data["assets"]["cash"] / data["liabilities"]["debt"]
    debt_equity = data["liabilities"]["debt"] / data["liabilities"]["equity"]
    roa = (data["net_profit"] / data["assets"]["total"]) * 100
    
    with col1:
        st.metric("Current Ratio", f"{current_ratio:.2f}", "Target: >1.5")
    with col2:
        st.st.metric("Debt/Equity", f"{debt_equity:.2f}", "Target: <1.0")
    with col3:
        st.metric("ROA", f"{roa:.1f}%", "Target: >15%")
    
    # Shark Verdict
    st.subheader("🦈 Shark Verdict")
    if current_ratio > 1.5 and debt_equity < 1.0 and roa > 15:
        st.markdown('<div class="shark-win"><h2>🏆 DEAL CLOSED! ₹1 Crore Investment! 🎉</h2></div>', unsafe_allow_html=True)
        st.balloons()
    else:
        st.markdown('<div class="shark-lose"><h3>💸 OUT! Improve your financials!</h3></div>', unsafe_allow_html=True)
        
        st.info("💡 **Tips to Win:**\n- Increase cash reserves\n- Reduce debt levels\n- Boost profitability")

elif game_mode == "Excel Template":
    st.header("📥 Download Excel Template")
    
    # Create Excel workbook
    wb = Workbook()
    ws_cf = wb.active
    ws_cf.title = "Cash Flows"
    
    # Cash Flow Sheet
    ws_cf['A1'] = "🦈 Finance Shark Game - MSRUAS"
    ws_cf['A1'].font = Font(bold=True, size=16)
    
    ws_cf['A3'] = "Discount Rate (%)"
    ws_cf['B3'] = 10
    
    headers = ["Year", "Cash Flow", "Discounted CF"]
    for col, header in enumerate(headers, 1):
        ws_cf.cell(row=5, column=col, value=header).font = Font(bold=True)
    
    sample_cf = [-5000000, 1000000, 2000000, 3500000, 5000000, 7000000]
    years = ["Y0 (Invest)", "Y1", "Y2", "Y3", "Y4", "Y5"]
    
    for i, (year, cf) in enumerate(zip(years, sample_cf)):
        ws_cf[f'A{7+i}'] = year
        ws_cf[f'B{7+i}'] = cf
        ws_cf[f'C{7+i}'] = f"=B{7+i}/(1+$B$3/100)^{i}"
    
    ws_cf['A16'] = "NPV"
    ws_cf['C16'] = "=SUM(C7:C12)"
    
    # Balance Sheet Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs['A1'] = "Balance Sheet Analysis"
    ws_bs['A1'].font = Font(bold=True, size=16)
    
    bs_items = ["Cash", "Inventory", "Total Assets", "Debt", "Equity"]
    bs_values = [1500000, 2000000, "=SUM(B2:B3)+3500000", 3000000, 5500000]
    
    for i, (item, val) in enumerate(zip(bs_items, bs_values), 2):
        ws_bs[f'A{i}'] = item
        ws_bs[f'B{i}'] = val
    
    # Ratios
    ws_bs['D1'] = "Ratios"
    ws_bs['D2'] = "Current Ratio"
    ws_bs['E2'] = "=B2/B5"
    ws_bs['D3'] = "Debt/Equity"
    ws_bs['E3'] = "=B5/B6"
    ws_bs['D4'] = "ROA %"
    ws_bs['E4'] = "=1200000/B4*100"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    st.download_button(
        label="📊 Download Financial Game Excel Template",
        data=output.getvalue(),
        file_name="finance_shark_game.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Footer
st.markdown("---")
st.markdown("*MSRUAS Management Fest 2026 | Theme: TV Series Business Analysis*")
